"""
엑셀 내보내기 모음.
  · build_match_excel()    상세보기 팝업 1경기를 한 시트로 (팝업과 같은 2단 배치)
  · build_upload_template() 경기 업로드용 빈 표본 양식
  · build_table_excel()    현재 조회된 분석표 그대로

모두 이미 계산·저장되어 있는 값을 옮겨 담을 뿐, 새로 계산하지 않는다.
"""
import io
import re

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 실제로 다 채운 한 경기 예시 — 작성안내 시트에 '이렇게 채우면 됩니다'로 보여준다.
# DT는 일부러 08/22 같은 실수하기 쉬운 형식이 아니라 정확한 형식(연도-월-일)으로 적어 둔다.
UPLOAD_EXAMPLE_ROW = {
    "L": "EP", "S": "25-26", "R": "1R", "No": 1, "DT": "2026-08-22", "TM": 1500,
    "HT": "아스널", "HS": 2, "RT": "핸승", "AS": 0, "AT": "첼시",
    "KW": 1.80, "KD": 3.50, "KL": 4.20, "KH": -1.0, "KHW": 2.10, "KHD": 3.60, "KHL": 1.80,
    "FW": 1.90, "FD": 3.60, "FL": 4.00, "FH": -1.0, "FHW": 2.00, "FHD": 3.50, "FHL": 1.85,
}

# 업로드 양식의 컬럼 — 사람이 엑셀로 직접 채우는 칸이다.
# (여기 순서대로 헤더를 만들어야 업로드 시 그대로 인식됨)
UPLOAD_TEMPLATE_COLS = [
    "L", "S", "R", "No", "DT", "TM", "HT", "HS", "RT", "AS", "AT",
    "KW", "KD", "KL", "KH", "KHW", "KHD", "KHL",
    "FW", "FD", "FL", "FH", "FHW", "FHD", "FHL",
]

# 최종배당(배변이 다 반영된 값) — 초기배당과 짝을 이루는 같은 칸을 E 접두사로 둔다.
#   국내(EK*) : 와이즈토토 배당변경 이력의 '지금 값'
#   해외(EF*) : 스코어맨 Bet365 '라이브' 배당
# 업로드 양식에는 일부러 넣지 않는다 — 크롤링으로만 나오는 값이라 사람이 손으로 채울
# 일이 없고, 양식만 넓어져 헷갈리기 때문이다.
FINAL_ODDS_COLS_K = ["EKW", "EKD", "EKL", "EKH", "EKHW", "EKHD", "EKHL"]
FINAL_ODDS_COLS_F = ["EFW", "EFD", "EFL", "EFH", "EFHW", "EFHD", "EFHL"]
FINAL_ODDS_COLS = FINAL_ODDS_COLS_K + FINAL_ODDS_COLS_F

# '계산해서 만든 값'이 아니라 원본으로 들어오는 칸 전부.
# engine.preprocess_data()의 target_cols와 내용이 같아야 한다 — 저장 경로(엑셀 업로드·
# 스코어맨 크롤·국배 크롤)가 전부 이 목록을 기준으로 원본/분석 컬럼을 가른다.
RAW_DATA_COLS = UPLOAD_TEMPLATE_COLS + FINAL_ODDS_COLS

# 경기입력 시트 맨 위에 놓는 한글 라벨(짧게). 이 줄은 영문 코드 헤더보다 위에 있어야 안전하다 —
# find_header_row()는 'DT'와 'HT'가 '같이 있는 줄'을 찾아 그걸 진짜 헤더로 삼으므로,
# 한글 라벨 줄이 코드 헤더 '위'에 있으면 자동으로 무시되고, '아래'에 있으면 데이터로 잘못 읽힌다.
UPLOAD_COL_KOREAN_LABEL = {
    "L": "리그", "S": "시즌", "R": "라운드", "No": "경기순서", "DT": "경기일", "TM": "경기시간",
    "HT": "홈팀", "HS": "홈팀점수", "RT": "결과", "AS": "원정팀 점수", "AT": "원정팀",
    "KW": "국)승", "KD": "국)무", "KL": "국)패", "KH": "국핸디",
    "KHW": "국)H-승", "KHD": "국)H-무", "KHL": "국)H-패",
    "FW": "해)승", "FD": "해)무", "FL": "해)패", "FH": "해핸디",
    "FHW": "해)H-승", "FHD": "해)H-무", "FHL": "해)H-패",
    # 최종배당(배변 후) — 업로드 양식에는 안 들어가고 분석표 다운로드에서만 쓰인다.
    "EKW": "국)승-최종", "EKD": "국)무-최종", "EKL": "국)패-최종", "EKH": "국핸디-최종",
    "EKHW": "국)H-승-최종", "EKHD": "국)H-무-최종", "EKHL": "국)H-패-최종",
    "EFW": "해)승-최종", "EFD": "해)무-최종", "EFL": "해)패-최종", "EFH": "해핸디-최종",
    "EFHW": "해)H-승-최종", "EFHD": "해)H-무-최종", "EFHL": "해)H-패-최종",
}

# 업로드 양식 각 컬럼의 뜻 — '작성안내' 시트 표에 함께 넣어 비개발자도 채울 수 있게 한다.
UPLOAD_COL_HINTS = {
    "L": "리그", "S": "시즌(예: 25-26)", "R": "라운드(예: 38R)", "No": "경기번호",
    "DT": "일자 (반드시 2026-08-22 처럼 '연도-월-일' 형식. 08/22처럼 연도 없이 적으면 인식 안 됨)",
    "TM": "시각", "HT": "홈팀(필수)", "HS": "홈 득점",
    "RT": "결과(핸승/핸무/무/역·취소·연기·미정이면 공란)", "AS": "원정 득점", "AT": "원정팀(필수)",
    "KW": "국내 홈", "KD": "국내 무", "KL": "국내 원정", "KH": "국내 핸디",
    "KHW": "국내핸디 홈", "KHD": "국내핸디 무", "KHL": "국내핸디 원정",
    "FW": "해외 홈", "FD": "해외 무", "FL": "해외 원정", "FH": "해외 핸디",
    "FHW": "해외핸디 홈", "FHD": "해외핸디 무", "FHL": "해외핸디 원정",
}

RT_LABELS = {1: "핸승", 2: "핸무", 3: "무", 4: "역", 5: "취소", 6: "연기"}
# DT 안 영문 요일(3글자) → 한글 한 글자. web/src/utils/format.js formatDt와 동일.
_WEEKDAY_KO = {"Sun": "일", "Mon": "월", "Tue": "화", "Wed": "수", "Thu": "목", "Fri": "금", "Sat": "토"}

# MatchDetailModal.jsx SAMPLE_INDICATORS와 정확히 같은 순서·라벨(K-PL="27. 국) 플핸
# 분석" 포함 21개) — 예전엔 이 목록이 K-PL 없이 20개짜리로 뒤처져 있었다.
SAMPLE_INDICATORS = [
    ("K-W", "국) 승"), ("K-L", "국) 패"), ("K-PL", "국) 플핸"),
    ("K-WL", "국) 승+패"), ("K-WDL", "국) 승+무+패"),
    ("K-W-HT", "국) 승=홈팀"), ("K-L-AT", "국) 패=원정팀"),
    ("TK-W", "국/통) 승"), ("TK-L", "국/통) 패"), ("TK-WL", "국/통) 승+패"), ("TK-WDL", "국/통) 승+무+패"),
    ("F-W", "해) 승"), ("F-L", "해) 패"), ("F-WL", "해) 승+패"), ("F-WDL", "해) 승+무+패"),
    ("F-W-HT", "해) 승=홈팀"), ("F-L-AT", "해) 패=원정팀"),
    ("TF-W", "해/통) 승"), ("TF-L", "해/통) 패"), ("TF-WL", "해/통) 승+패"), ("TF-WDL", "해/통) 승+무+패"),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_CENTER = Alignment(horizontal="center", vertical="center")
_THIN = Side(style="thin", color="D1D5DB")
_THICK = Side(style="medium", color="6B7280")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_DIVIDER_BORDER = Border(left=_THIN, right=_THIN, top=_THICK, bottom=_THIN)  # 화면의 굵은 구분선과 동일
_TITLE_FONT = Font(bold=True, size=14)
_BOLD = Font(bold=True)


def _with_right_divider(border):
    """기존 테두리는 유지한 채 오른쪽만 굵게 — 역/토탈 사이 세로 구분선용."""
    b = border or _BORDER
    return Border(left=b.left, right=_THICK, top=b.top, bottom=b.bottom)
_WINNER_FONT = Font(color="C62828", bold=True)  # 화면의 이긴 팀 점수 빨간 강조와 동일


def _rt_label(v):
    if v is None or v == "":
        return None
    try:
        return RT_LABELS.get(int(float(v)))
    except (TypeError, ValueError):
        return None


def _num_or_dash(v, digits=2):
    if v is None or v == "":
        return "-"
    try:
        return round(float(v), digits)
    except (TypeError, ValueError):
        return "-"


def _pct_or_dash(v):
    if v is None or v == "":
        return "-"
    try:
        return f"{float(v):.0f}%"
    except (TypeError, ValueError):
        return "-"


def _form_or_dash(v):
    """폼(PPG) 값은 백엔드가 '2.13' 같은 문자열로 이미 반올림해 보내준다 — 그대로 쓴다."""
    return "-" if v is None or v == "" else str(v)


_HOME_LETTER_FONT = InlineFont(b=True, u="single")


def _spaced_with_home_underline(results, venues):
    """results('WWDDL')와 venues('HAHAA', 같은 자리수의 H/A)를 같이 훑어서
    화면의 '점 = 홈경기' 표시를 엑셀에서는 그 글자에 밑줄로 표시한다.
    (결합 문자(dot above) 방식은 엑셀/폰트에 따라 점이 잘 안 보여서 밑줄로 바꿨다)

    글자 사이는 칸 2개로 띄우되, 그 공백을 별도 런(run)으로 두지 않고 글자 뒤에
    붙여 한 런으로 만든다 — openpyxl의 xml:space="preserve" 처리(whitespace())가
    "공백만 있는 런"은 못 알아채고 건너뛰어서, 공백만 담은 런을 따로 두면 진짜
    엑셀에서 열었을 때 그 공백이 통째로 사라져 글자가 다 붙어 보인다."""
    if not results:
        return "-"
    venues = venues or ""
    n = len(results)
    parts = []
    for i, ch in enumerate(results):
        is_home = i < len(venues) and venues[i] == "H"
        text = ch if i == n - 1 else ch + "  "
        parts.append(TextBlock(_HOME_LETTER_FONT, text) if is_home else text)
    return CellRichText(*parts)


def _int_or_blank(v):
    if v is None or v == "":
        return ""
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return v


_RT_ORDER = ["핸승", "핸무", "무", "역"]


def _wdl_breakdown_text(bucket):
    """HeadToHeadResult.jsx wdlBreakdownText()와 동일 — "핸승(2) / 핸무(3) / 역(1)" 형태."""
    if not bucket:
        return ""
    entries = list((bucket.get("breakdown") or {}).items())
    entries.sort(key=lambda kv: _RT_ORDER.index(kv[0]) if kv[0] in _RT_ORDER else len(_RT_ORDER))
    return " / ".join(f"{label}({n})" for label, n in entries)


def _points_for(m, reference_team):
    """HeadToHeadResult.jsx homePoints()와 동일 — 기준 팀이 홈/원정 상관없이 그 경기에서
    딴 승점(승3/무1/패0). 점수가 없으면(예정 경기) 빈 칸."""
    hs, as_ = m.get("HS"), m.get("AS")
    if hs is None or as_ is None or hs == "" or as_ == "":
        return ""
    try:
        hs, as_ = float(hs), float(as_)
    except (TypeError, ValueError):
        return ""
    if m.get("HT") == reference_team:
        mine, theirs = hs, as_
    elif m.get("AT") == reference_team:
        mine, theirs = as_, hs
    else:
        return ""
    if mine > theirs:
        return 3
    if mine < theirs:
        return 0
    return 1


_RIGHT_COL = 8  # 'H' — 팝업의 오른쪽 칼럼(시즌전적 등)이 시작하는 위치.
# 왼쪽(지표별 표본 등)이 F열까지 쓰므로, G열은 비워 두 블록 사이에 한 칸 여백을 둔다.

# 화면(RiskCard riskCellStyle)과 같은 5단계 색 — "초록이면 플핸에 유리".
_R_DEEP = {"bg": "1B5E20", "fg": "FFFFFF", "bold": True}
_R_GOOD = {"bg": "66BB6A", "fg": "0D1B2A", "bold": True}
_R_MID = {"bg": "FBC02D", "fg": "0D1B2A", "bold": False}
_R_WARN = {"bg": "EF6C00", "fg": "FFFFFF", "bold": True}
_R_BAD = {"bg": "C62828", "fg": "FFFFFF", "bold": True}


def _risk_style(kind, n):
    """확률 지표(정승%/플핸무%/플%) 칸 색 — MatchDetailModal.jsx riskCellStyle과 동일 경계."""
    if n is None:
        return None
    if kind == "win":      # 정승 — 정배가 셀수록 플핸에 불리
        if n < 40:
            return _R_GOOD
        if n < 55:
            return _R_MID
        if n < 70:
            return _R_WARN
        return _R_BAD
    if kind == "nh":        # 플핸무 — 실측 평균 68~70%
        if n >= 85:
            return _R_DEEP
        if n >= 75:
            return _R_GOOD
        if n >= 65:
            return _R_MID
        if n >= 55:
            return _R_WARN
        return _R_BAD
    # pl — 플, 실측 평균 44~46%
    if n >= 55:
        return _R_DEEP
    if n >= 48:
        return _R_GOOD
    if n >= 41:
        return _R_MID
    if n >= 34:
        return _R_WARN
    return _R_BAD


_DDONG_GRADES = [(22, "안전"), (30, "보통"), (37, "주의")]


def _ddong_grade(risk):
    for cut, label in _DDONG_GRADES:
        if risk < cut:
            return label
    return "위험"


def _home_is_fav(row):
    """정배(시장이 강하다고 본 쪽)가 홈인지 — MatchDetailModal.jsx homeIsFav와 동일
    우선순위(국내배당 KW/KL 우선, 없으면 해외배당 FW/FL)."""
    for wk, lk in (("KW", "KL"), ("FW", "FL")):
        w, l = _num(row.get(wk)), _num(row.get(lk))
        if w is not None and l is not None and w != l:
            return w < l
    return None


def _fav_sample_codes(row):
    """지표별 표본에서 '판단에 쓰는 지표' 코드 집합 — MatchDetailModal.jsx
    favSampleCodes와 동일 규칙(정배 방향에 따라 갈리는 4줄 + 방향 무관 4줄 + K-PL)."""
    out = {"F-WL", "F-WDL", "K-WL", "K-WDL", "K-PL"}

    def pick(wk, lk, wcode, lcode, home_win_code, away_lose_code):
        w, l = _num(row.get(wk)), _num(row.get(lk))
        if w is None or l is None or w == l:
            return
        if w < l:
            out.add(wcode)
            out.add(home_win_code)
        else:
            out.add(lcode)
            out.add(away_lose_code)

    pick("KW", "KL", "K-W", "K-L", "K-W-HT", "K-L-AT")
    pick("FW", "FL", "F-W", "F-L", "F-W-HT", "F-L-AT")
    return out


def _max_second(vals):
    """각 값에 'max'/'second'/None — MatchDetailModal.jsx maxCellClass와 동일 규칙
    (1등·2등이 서로 다른 값일 때만 2등도 표시, 전부 0이면 강조 없음)."""
    m = max(vals) if vals else 0
    if m <= 0:
        return [None] * len(vals)
    smaller = [v for v in vals if v < m]
    second = max(smaller) if smaller else 0
    out = []
    for v in vals:
        if v == m:
            out.append("max")
        elif second > 0 and v == second:
            out.append("second")
        else:
            out.append(None)
    return out


_MAX_FILL2 = PatternFill("solid", fgColor="FDE68A")     # 화면 cell-max와 같은 계열(진하게)
_SECOND_FILL = PatternFill("solid", fgColor="FEF3C7")    # 화면 cell-second(옅게)
_FAV_SIDE = Side(style="medium", color="F59E0B")          # 화면 --ddong-focus 강조 테두리
_FAV_BORDER = Border(left=_FAV_SIDE, right=_FAV_SIDE, top=_FAV_SIDE, bottom=_FAV_SIDE)
_ODDS_FAV_FILL = PatternFill("solid", fgColor="DCE9FA")   # 화면 odds-fav-col과 같은 옅은 파랑


def build_match_excel(row: dict, h2h: dict, scope: str = "master",
                      season_rows: list | None = None, streaks: dict | None = None,
                      ht_record: dict | None = None, at_record: dict | None = None) -> io.BytesIO:
    """
    MatchDetailModal.jsx 현재 화면을 그대로 옮긴다.
      확률 지표(정승%/플핸무%/플%, 전체 폭) → 왼쪽: 배당·지표별 표본(전체 21줄) /
      오른쪽: 시즌전적·폼 지표·최근10경기+연속기록·상대전적.
    row/h2h/season_rows/streaks/ht_record/at_record는 화면과 같은 계산 결과를 그대로
    받는다(main.py가 /api/pick_ai·team_bet_record와 같은 함수를 불러 넘겨준다) — 여기서
    새로 계산하지 않는다.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "상세보기"

    def write_row(values, row_idx, start_col=1, font=None, fill=None, align=None, border=None,
                 number_format=None):
        for i, v in enumerate(values):
            cell = ws.cell(row=row_idx, column=start_col + i, value=v)
            cell.border = border or _BORDER
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if align:
                cell.alignment = align
            if number_format:
                cell.number_format = number_format
        return row_idx + 1

    ht = str(row.get("HT") or "").strip()
    at = str(row.get("AT") or "").strip()
    rt = _rt_label(row.get("RT"))
    hs, as_ = row.get("HS"), row.get("AS")
    has_score = hs is not None and as_ is not None and hs != "" and as_ != ""
    home_fav = _home_is_fav(row)

    def rank_suffix(v):
        n = _num(v)
        return f"({int(n)}위)" if n is not None else ""

    def role_suffix(is_home):
        if home_fav is None:
            return ""
        is_fav = home_fav if is_home else not home_fav
        return " (정)" if is_fav else " (역)"

    def bet_suffix(rec):
        if not rec or not rec.get("total"):
            return ""
        return f" ({rec['hit']}/{rec['total']})"

    r = 1
    star = "⭐ " if row.get("IMPORTANT") else ""
    title = (f"{star}{ht}{rank_suffix(row.get('HP'))}{role_suffix(True)}{bet_suffix(ht_record)}"
             f" vs {at}{rank_suffix(row.get('AP'))}{role_suffix(False)}{bet_suffix(at_record)}")
    ws.cell(row=r, column=1, value=title).font = _TITLE_FONT
    r += 1

    # DT는 'YY-MM-DD (Sat)'처럼 요일이 영문 3글자로 저장돼 있다 — 화면(formatDt)과 같은
    # 한글 한 글자로 바꿔서 보여준다. TM(HHMM 숫자)도 화면(formatTime)처럼 'HH:MM'으로.
    dt_txt = re.sub(r"\(([A-Za-z]{3})\)", lambda m: f"({_WEEKDAY_KO.get(m.group(1), m.group(1))})",
                    str(row.get("DT")) if row.get("DT") else "")
    tm_val = _num(row.get("TM"))
    tm_txt = f"{int(tm_val):04d}"[:2] + ":" + f"{int(tm_val):04d}"[2:] if tm_val is not None else ""
    meta = f"{row.get('S', '')} · {row.get('R', '')}"
    if dt_txt:
        meta += f" · {dt_txt}"
    if tm_txt:
        meta += f" {tm_txt}"
    meta += f" · {rt}" if rt else " · 예정 경기"
    ws.cell(row=r, column=1, value=meta)
    r += 1

    # 내픽/P/의견/메모 — 화면 상단 MyPickBar와 같은 내용(계정 개인 기록).
    pick_bits = []
    if row.get("MY_PICK"):
        pick_bits.append(f"내픽 {row['MY_PICK']}")
    if row.get("MY_P"):
        pick_bits.append(f"P {row['MY_P']}")
    if row.get("MY_HIT"):
        pick_bits.append(f"의견 {row['MY_HIT']}")
    if pick_bits:
        ws.cell(row=r, column=1, value=" · ".join(pick_bits))
        r += 1
    if row.get("MEMO"):
        ws.cell(row=r, column=1, value=f"메모: {row['MEMO']}")
        r += 1

    if has_score:
        hs_val, as_val = _int_or_blank(hs), _int_or_blank(as_)
        # 이긴 팀 점수만 빨간 강조(무승부는 그대로) — 화면의 winner-score와 동일 규칙.
        # 셀 하나에 다 넣으면 부분 강조가 안 되어 홈/원정/점수/구분자를 각 칸에 나눠 적는다.
        r = write_row([ht, hs_val, ":", as_val, at], r, font=_BOLD)
        if isinstance(hs_val, int) and isinstance(as_val, int):
            if hs_val > as_val:
                ws.cell(row=r - 1, column=2).font = _WINNER_FONT
            elif as_val > hs_val:
                ws.cell(row=r - 1, column=4).font = _WINNER_FONT
    r += 1

    # ── 확률 지표(RiskCard, 전체 폭) — 정승%/플핸무%/플% 8칸을 화면과 같은 순서로 ──
    ws.cell(row=r, column=1, value="확률 지표").font = _BOLD
    r += 1
    risk_groups = [
        ("정승 %", "win", [("국)정", row.get("WIN_RISK")), ("해)정", row.get("WIN_RISK_F"))]),
        ("플핸무 %", "nh", [("국)플", row.get("NH_KO")), ("국)지", row.get("NH_KI")),
                          ("해)지", row.get("NH_FI"))]),
        ("플 %", "pl", [("국)플", row.get("PL_KO")), ("국)지", row.get("PL_KI")),
                       ("해)지", row.get("PL_FI"))]),
    ]
    head1 = r
    col = 1
    for title_g, _kind, cols in risk_groups:
        span = len(cols)
        c = ws.cell(row=head1, column=col, value=title_g)
        c.font, c.fill, c.alignment, c.border = _HEADER_FONT, _HEADER_FILL, _CENTER, _BORDER
        if span > 1:
            ws.merge_cells(start_row=head1, start_column=col, end_row=head1, end_column=col + span - 1)
            for i in range(1, span):
                ws.cell(row=head1, column=col + i).border = _BORDER
        col += span
    r += 1
    head2 = r
    col = 1
    for _title_g, _kind, cols in risk_groups:
        for label, _v in cols:
            c = ws.cell(row=head2, column=col, value=label)
            c.font, c.fill, c.alignment, c.border = _HEADER_FONT, _HEADER_FILL, _CENTER, _BORDER
            col += 1
    r += 1
    data_row = r
    col = 1
    for _title_g, kind, cols in risk_groups:
        for _label, val in cols:
            n = _num(val)
            c = ws.cell(row=data_row, column=col, value=f"{n:.0f}%" if n is not None else "-")
            c.alignment, c.border = _CENTER, _BORDER
            style = _risk_style(kind, n)
            if style:
                c.fill = PatternFill("solid", fgColor=style["bg"])
                c.font = Font(bold=style["bold"], color=style["fg"])
            col += 1
    r += 2
    section_start = r  # 배당/시즌전적이 나란히 시작하는 행

    # ── 왼쪽: 배당 → 지표별 표본(전체 21줄) ──
    ws.cell(row=r, column=1, value="배당").font = _BOLD
    ddong = str(row.get("DDONG") or "").strip()
    if ddong:
        ddong_risk = _num(row.get("DDONG_RISK"))
        note = ddong
        if ddong_risk is not None:
            note += f" · {_ddong_grade(ddong_risk)} {ddong_risk:.0f}%"
        if str(row.get("DDONGSA") or "").strip():
            note += " · 똥사"
        ws.cell(row=r, column=2, value=note)
    r += 1
    r = write_row(["구분", "승(홈)", "무", "패(원정)"], r, font=_HEADER_FONT,
                  fill=_HEADER_FILL, align=_CENTER)
    # 정배 쪽 칸(승=홈팀 칸 / 패=원정팀 칸)에 화면(odds-fav-col)과 같은 옅은 파랑 배경.
    # 핸디 배당(국내 핸디·해외 핸디) 줄만은 정배가 아니라 핸디를 받은 언더독 쪽을
    # 강조한다 — 핸디 시장에서 보는 값은 "언더독이 그 핸디를 커버하는가"이므로 늘
    # 언더독 칸이 관심 대상이다(화면 OddsTable의 dogColClass와 동일 규칙).
    fav_col = "w" if home_fav is True else ("l" if home_fav is False else None)
    dog_col = "l" if home_fav is True else ("w" if home_fav is False else None)
    for label, w, d, l, is_handi in (
        ("국내 배당", "KW", "KD", "KL", False),
        ("국내 핸디", "KHW", "KHD", "KHL", True),
        ("해외 배당", "FW", "FD", "FL", False),
        ("해외 핸디", "FHW", "FHD", "FHL", True),
    ):
        this_fav = dog_col if is_handi else fav_col
        row_idx = r
        # 숫자를 그대로 넣으면 4.80처럼 끝의 0이 엑셀 기본서식(General)에서 사라져 4.8로
        # 보인다 — 화면 팝업(toFixed(2))과 같은 자리수로 보이게 셀 서식을 소수 둘째자리로 고정.
        r = write_row([label, _num_or_dash(row.get(w)), _num_or_dash(row.get(d)),
                       _num_or_dash(row.get(l))], r, number_format="0.00")
        if this_fav == "w":
            ws.cell(row=row_idx, column=2).fill = _ODDS_FAV_FILL
        elif this_fav == "l":
            ws.cell(row=row_idx, column=4).fill = _ODDS_FAV_FILL
    r += 1

    ws.cell(row=r, column=1, value="지표별 표본").font = _BOLD
    r += 1
    _SAMPLE_YK_COL = 5  # 지표(1) 핸승(2) 핸무(3) 무(4) 역(5) 토탈(6)
    r = write_row(["지표", "핸승", "핸무", "무", "역", "토탈"], r, font=_HEADER_FONT,
                  fill=_HEADER_FILL, align=_CENTER)
    ws.cell(row=r - 1, column=_SAMPLE_YK_COL).border = _with_right_divider(_BORDER)
    # 내 데이터(scope=user)는 통합(TK-/TF-) 지표를 뺀다 — 화면(SampleTable)과 동일:
    # 리그 하나만 있어 통합 대상이 없으므로 국내/해외 지표와 값이 완전히 같아지는
    # 의미 없는 중복이기 때문이다.
    indicators = SAMPLE_INDICATORS if scope != "user" else [
        (c, lb) for c, lb in SAMPLE_INDICATORS if not c.startswith("TK-") and not c.startswith("TF-")
    ]
    fav_codes = _fav_sample_codes(row)
    grand = [0, 0, 0, 0]
    prev_code = None
    for code, label in indicators:
        vals = []
        for i in (1, 2, 3, 4):
            try:
                vals.append(int(float(row.get(f"{code} {i}") or 0)))
            except (TypeError, ValueError):
                vals.append(0)
        for i, v in enumerate(vals):
            grand[i] += v
        total = sum(vals)
        row_idx = r
        # 국내(K-/TK-) 지표 블록 → 해외(F-/TF-) 지표 블록 경계에 화면과 같은 굵은 선
        is_foreign = code.startswith("F-") or code.startswith("TF-")
        was_foreign = prev_code is not None and (prev_code.startswith("F-") or prev_code.startswith("TF-"))
        group_border = _DIVIDER_BORDER if (is_foreign and not was_foreign) else None
        # 위=비율, 아래=건수를 한 칸에 — 화면(sample-n)과 같은 정보를 한 칸으로 압축.
        cells = [f"{round(v / total * 100)}% ({v})" if total > 0 else "-" for v in vals]
        r = write_row([label, *cells, total], r, border=group_border)
        ws.cell(row=row_idx, column=_SAMPLE_YK_COL).border = _with_right_divider(group_border)
        # 오늘 이 경기의 판단에 쓰는 지표(favSampleCodes)는 화면처럼 강조 테두리로 감싼다.
        if code in fav_codes:
            for cc in range(1, 7):
                ws.cell(row=row_idx, column=cc).border = _FAV_BORDER
        prev_code = code
        cls = _max_second(vals)
        for i, c in enumerate(cls):
            if c == "max":
                ws.cell(row=row_idx, column=2 + i).fill = _MAX_FILL2
            elif c == "second":
                ws.cell(row=row_idx, column=2 + i).fill = _SECOND_FILL

    row_idx = r
    gtotal = sum(grand)
    gcells = [f"{round(v / gtotal * 100)}% ({v})" if gtotal > 0 else "-" for v in grand]
    r = write_row(["토탈", *gcells, gtotal], r, font=_BOLD)
    ws.cell(row=row_idx, column=_SAMPLE_YK_COL).border = _with_right_divider(_BORDER)
    gcls = _max_second(grand)
    for i, c in enumerate(gcls):
        if c == "max":
            ws.cell(row=row_idx, column=2 + i).fill = _MAX_FILL2
        elif c == "second":
            ws.cell(row=row_idx, column=2 + i).fill = _SECOND_FILL

    # ── 오른쪽(H열): 시즌전적 → 폼 지표 → 최근10경기 전적+연속기록 → 상대전적 ──
    rr = section_start
    if season_rows:
        ws.cell(row=rr, column=_RIGHT_COL, value="시즌전적").font = _BOLD
        rr += 1
        ws.cell(row=rr, column=_RIGHT_COL,
               value="오늘과 같은 정배/역배 구도였던 이번 시즌 경기만 모은 값입니다.")
        rr += 1
        rr = write_row(["", "핸승", "핸무", "무", "역"], rr, start_col=_RIGHT_COL,
                       font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
        for sr in season_rows:
            side_label = sr.get("side") or ""
            role = sr.get("role")
            label = f"{side_label}({role})" if role else side_label
            counts = sr.get("counts") or {}
            rr = write_row([label, counts.get("핸승", "-"), counts.get("핸무", "-"),
                           counts.get("무", "-"), counts.get("역", "-")], rr, start_col=_RIGHT_COL)
        rr += 1

    ws.cell(row=rr, column=_RIGHT_COL, value="폼 지표").font = _BOLD
    rr += 1
    _FORM_DIV_COL = _RIGHT_COL + 2   # 홈 3칸 / 원정 3칸 경계
    form_head = rr
    rr = write_row(["홈", "", "", "원정", "", ""], rr, start_col=_RIGHT_COL,
                   font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    ws.merge_cells(start_row=form_head, start_column=_RIGHT_COL,
                   end_row=form_head, end_column=_RIGHT_COL + 2)
    ws.merge_cells(start_row=form_head, start_column=_RIGHT_COL + 3,
                   end_row=form_head, end_column=_RIGHT_COL + 5)
    ws.cell(row=form_head, column=_FORM_DIV_COL).border = _with_right_divider(_BORDER)
    for values in (
        ["전체폼", "최근5폼", "홈경기", "원정경기", "최근5폼", "전체폼"],
        [_form_or_dash(row.get(k)) for k in ("HTF", "HRF", "HF", "AF", "ARF", "ATF")],
    ):
        is_head = values[0] == "전체폼"
        line = rr
        rr = write_row(values, rr, start_col=_RIGHT_COL,
                       font=_HEADER_FONT if is_head else _BOLD,
                       fill=_HEADER_FILL if is_head else None, align=_CENTER)
        ws.cell(row=line, column=_FORM_DIV_COL).border = _with_right_divider(_BORDER)
    rr += 1

    ws.cell(row=rr, column=_RIGHT_COL, value="최근10경기 전적").font = _BOLD
    rr += 1
    head = rr
    rr = write_row(["홈팀최근 →", "", "", "← 원정팀 최근", "", ""], rr, start_col=_RIGHT_COL,
                   font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
    ws.merge_cells(start_row=head, start_column=_RIGHT_COL,
                   end_row=head, end_column=_RIGHT_COL + 2)
    ws.merge_cells(start_row=head, start_column=_RIGHT_COL + 3,
                   end_row=head, end_column=_RIGHT_COL + 5)
    ws.cell(row=head, column=_FORM_DIV_COL).border = _with_right_divider(_BORDER)
    line = rr
    # 홈팀은 왼쪽이 과거·오른쪽이 최신, 원정팀은 그 반대 (백엔드가 이미 그 순서로 만들어 둔다).
    # 밑줄 친 글자는 화면의 '그 경기가 홈경기였다' 표시와 동일하다.
    rr = write_row([
        _spaced_with_home_underline(row.get("HR10"), row.get("HR10H")), "", "",
        _spaced_with_home_underline(row.get("AR10"), row.get("AR10H")), "", "",
    ], rr, start_col=_RIGHT_COL, font=_BOLD, align=_CENTER)
    ws.merge_cells(start_row=line, start_column=_RIGHT_COL,
                   end_row=line, end_column=_RIGHT_COL + 2)
    ws.merge_cells(start_row=line, start_column=_RIGHT_COL + 3,
                   end_row=line, end_column=_RIGHT_COL + 5)
    ws.cell(row=line, column=_FORM_DIV_COL).border = _with_right_divider(_BORDER)
    rr += 1
    ws.cell(row=rr, column=_RIGHT_COL, value="밑줄 친 글자 = 그 팀 기준 홈경기")
    rr += 1

    # 최고 연속 기록(연승/무패/무승/연패) — 그 리그 데이터로만, 이 경기 직전까지.
    home_streak = (streaks or {}).get("home")
    away_streak = (streaks or {}).get("away")

    def streak_text(s):
        if not s or not s.get("played"):
            return ""
        return f"연승{s.get('win', 0)} · 무패{s.get('unbeaten', 0)} · 무승{s.get('winless', 0)} · 연패{s.get('lose', 0)}"

    home_txt, away_txt = streak_text(home_streak), streak_text(away_streak)
    if home_txt or away_txt:
        ws.cell(row=rr, column=_RIGHT_COL, value=home_txt)
        ws.merge_cells(start_row=rr, start_column=_RIGHT_COL, end_row=rr, end_column=_RIGHT_COL + 2)
        ws.cell(row=rr, column=_RIGHT_COL + 3, value=away_txt)
        ws.merge_cells(start_row=rr, start_column=_RIGHT_COL + 3, end_row=rr, end_column=_RIGHT_COL + 5)
        rr += 1

    ws.cell(row=rr, column=_RIGHT_COL, value="상대전적").font = _BOLD
    rr += 1
    ws.cell(row=rr, column=_RIGHT_COL,
           value=f"{ht} 기준 실제 승/무/패이며, 괄호는 그때 핸디캡 결과(RT)입니다.")
    rr += 1
    summary = h2h.get("summary")
    wdl = h2h.get("wdl_summary")
    if summary:
        _H2H_YK_COL = _RIGHT_COL + 2  # W(+0) D(+1) L(+2) 토탈(+3)
        if wdl:
            w_total, d_total, l_total = wdl["W"]["total"], wdl["D"]["total"], wdl["L"]["total"]
            rr = write_row([f"W({w_total})", f"D({d_total})", f"L({l_total})", "토탈"], rr,
                           start_col=_RIGHT_COL, font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
            ws.cell(row=rr - 1, column=_H2H_YK_COL).border = _with_right_divider(_BORDER)
            summary_row = rr
            rr = write_row([
                _wdl_breakdown_text(wdl["W"]), _wdl_breakdown_text(wdl["D"]),
                _wdl_breakdown_text(wdl["L"]), w_total + d_total + l_total,
            ], rr, start_col=_RIGHT_COL)
            ws.cell(row=summary_row, column=_H2H_YK_COL).border = _with_right_divider(_BORDER)
            rr += 1
        rr = write_row(["시즌", "R", "HT", "HS", "AS", "AT", "결과", "유형", "승점"], rr,
                       start_col=_RIGHT_COL, font=_HEADER_FONT, fill=_HEADER_FILL, align=_CENTER)
        prev_season = None
        _WDL_LETTER = {3: "W", 1: "D", 0: "L"}
        for m in h2h.get("matches", []):
            season = m.get("S")
            # 시즌이 바뀌는 경계마다 화면과 같은 굵은 구분선
            divider = prev_season is not None and season != prev_season
            hs_val = _int_or_blank(m.get("HS"))
            as_val = _int_or_blank(m.get("AS"))
            row_idx = rr
            points = _points_for(m, ht)
            rr = write_row([
                season, m.get("R"), m.get("HT"), hs_val, as_val, m.get("AT"),
                _WDL_LETTER.get(points, ""), m.get("RT_label") or "", points,
            ], rr, start_col=_RIGHT_COL, border=_DIVIDER_BORDER if divider else None)
            # 이긴 팀 점수만 빨간 강조 (화면의 winner-score와 동일)
            if isinstance(hs_val, int) and isinstance(as_val, int):
                if hs_val > as_val:
                    ws.cell(row=row_idx, column=_RIGHT_COL + 3).font = _WINNER_FONT
                elif as_val > hs_val:
                    ws.cell(row=row_idx, column=_RIGHT_COL + 4).font = _WINNER_FONT
            prev_season = season
        total = h2h.get("total", 0)
        shown = len(h2h.get("matches", []))
        if total > shown:
            ws.cell(row=rr, column=_RIGHT_COL, value=f"최근 {shown}경기만 표시 (총 {total}경기)")
            rr += 1
        ws.cell(row=rr, column=_RIGHT_COL, value="승점은 홈팀 기준으로 작성되었습니다.")
        rr += 1
    else:
        ws.cell(row=rr, column=_RIGHT_COL, value=f"{ht} vs {at} 맞대결 기록 없음")
        rr += 1

    widths = {1: 18, 2: 13, 3: 13, 4: 13, 5: 13, 6: 10,
             7: 4, 8: 10, 9: 10, 10: 10, 11: 10, 12: 10, 13: 10, 14: 10, 15: 10, 16: 10}
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


_KOREAN_LABEL_FONT = Font(bold=False, color="1A1A1A")


def build_upload_template(league_code: str = "", season: str = "", round_label: str = "",
                          match_count: int = 10) -> io.BytesIO:
    """
    경기 업로드용 빈 표본 양식.

    첫 시트('경기입력')는 2행 헤더 — 1행은 한글 라벨(리그/시즌/...), 2행은 실제
    인식되는 영문 코드(L/S/...)이고, 데이터는 3행부터 시작한다.
    한글 라벨을 영문 코드 '아래'에 넣으면(예전 방식) find_header_row()가 2행을
    헤더로 잡은 뒤 그 아래 한글 라벨 줄이 HT/AT가 채워진 경기 한 건으로 오인식되어
    쓰레기 데이터가 들어갔다. 한글 라벨을 코드 '위'에 두면 find_header_row()가
    'DT'+'HT'가 있는 2행을 정확히 헤더로 찾아내고, 그 위 1행은 자동으로 무시되어 안전하다.

    league_code(L)/season(S)/round_label(R)을 넘기면 경기번호(No) 1~match_count로
    행을 미리 만들어 그 세 컬럼을 채워 둔다 — 나머지 칸(HT/AT/배당 등)만 채우면 되게.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "경기입력"

    for i, col in enumerate(UPLOAD_TEMPLATE_COLS, start=1):
        label_cell = ws.cell(row=1, column=i, value=UPLOAD_COL_KOREAN_LABEL.get(col, col))
        label_cell.font = _KOREAN_LABEL_FONT
        label_cell.alignment = _CENTER
        label_cell.border = _BORDER
        if col == "DT":
            label_cell.comment = Comment(
                "일자는 반드시 연도-월-일 형식으로 적으세요.\n"
                "예: 2026-08-22\n"
                "08/22 처럼 연도 없이 적으면 인식되지 않습니다.",
                "BETPRO",
            )

        code_cell = ws.cell(row=2, column=i, value=col)
        code_cell.font = _HEADER_FONT
        code_cell.fill = _HEADER_FILL
        code_cell.alignment = _CENTER
        code_cell.border = _BORDER

        ws.column_dimensions[get_column_letter(i)].width = 14

    if league_code or season or round_label:
        prefill = {"L": league_code, "S": season, "R": round_label}
        for row_i in range(match_count):
            for col_i, col in enumerate(UPLOAD_TEMPLATE_COLS, start=1):
                if col == "No":
                    ws.cell(row=3 + row_i, column=col_i, value=row_i + 1)
                elif col in prefill and prefill[col]:
                    ws.cell(row=3 + row_i, column=col_i, value=prefill[col])

    ws.freeze_panes = "A3"

    guide = wb.create_sheet("작성안내")
    guide.cell(row=1, column=1, value="경기 데이터 작성 안내").font = _TITLE_FONT
    notes = [
        "'경기입력' 시트의 2행부터 한 줄에 한 경기씩 입력하세요.",
        "홈팀(HT)과 원정팀(AT)은 반드시 채워야 합니다. 비어 있으면 그 줄은 무시됩니다.",
        "결과(RT)는 핸승 / 핸무 / 무 / 역 중 하나로 적고, 아직 안 끝난 경기는 비워 두세요."
        " 아예 열리지 않은 경기는 '취소', 날짜만 미뤄진 경기는 '연기'로 적으면 됩니다"
        " (둘 다 통계에서 빠지며, 연기 경기는 나중에 치러지면 실제 결과로 고쳐 주세요).",
        "이미 등록된 경기와 시즌·라운드·경기번호·팀이 모두 같으면 새로 올린 값으로 대체됩니다"
        " (예정 경기에 결과만 나중에 채우는 용도). 단, 그 경기의 26개 지표·플핸예측은 처음"
        " 등록됐을 때 값을 그대로 유지하며 다시 계산되지 않습니다.",
        "컬럼 순서를 바꾸거나 컬럼명을 고치지 마세요.",
        "일자(DT)는 반드시 '연도-월-일' 형식으로 적으세요. 예: 2026-08-22"
        " (08/22처럼 연도 없이 적으면 인식되지 않습니다.)",
    ]
    r = 3
    for n in notes:
        guide.cell(row=r, column=1, value=f"· {n}")
        r += 1
    r += 2

    guide.cell(row=r, column=1, value="입력 예시 (한 경기를 다 채우면 이런 모습)").font = _BOLD
    r += 1
    for col_i, col in enumerate(UPLOAD_TEMPLATE_COLS, start=1):
        c = guide.cell(row=r, column=col_i, value=col)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
    r += 1
    for col_i, col in enumerate(UPLOAD_TEMPLATE_COLS, start=1):
        guide.cell(row=r, column=col_i, value=UPLOAD_EXAMPLE_ROW.get(col))
    r += 2

    guide.cell(row=r, column=1, value="컬럼").font = _BOLD
    guide.cell(row=r, column=2, value="설명").font = _BOLD
    r += 1
    for col in UPLOAD_TEMPLATE_COLS:
        guide.cell(row=r, column=1, value=col).font = _BOLD
        guide.cell(row=r, column=2, value=UPLOAD_COL_HINTS.get(col, ""))
        r += 1
    guide.column_dimensions["A"].width = 10
    guide.column_dimensions["B"].width = 42

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ════════════════════════════════════════════════════════════
# 분석표 2단 헤더 + 색상 — web/src/components/LeagueTable/columnGroups.js를
# 그대로 옮긴 것. 그룹 구성·라벨·색상 규칙이 어긋나면 안 되므로 순서·값 모두 동일하게 유지.
# ════════════════════════════════════════════════════════════
_GEN_COLS = ["L", "S", "R", "No", "DT", "TM"]
# 경기 직전 시즌 성적: HP/AP=순위, HTF/ATF=전체경기 PPG, HF/AF=홈·원정경기 PPG
_MATCH_COLS = ["HTF", "HF", "HP", "HT", "HS", "RT", "AS", "AT", "AP", "AF", "ATF"]
_K_ODDS_COLS = ["KW", "KD", "KL", "KH", "KHW", "KHD", "KHL"]
_F_ODDS_COLS = ["FW", "FD", "FL", "FH", "FHW", "FHD", "FHL"]
_PH_COLS = [
    ("PH_F", "해)플핸"), ("PH_K", "국)플핸"), ("PH_PICK", "PICK"),
    ("PH_HIT", "실측"), ("PH_DOM", "비중"),
]
_GROUP_DEFS = [
    ("F-W", "1. 해) 승 분석"), ("F-L", "2. 해) 패 분석"), ("F-WL", "3. 해) 승+패 분석"),
    ("F-WDL", "4. 해) 승+무+패 분석"), ("F-W-HW", "5. 해) 승+H핸 분석"),
    ("F-W-HT", "6. 해) 승=홈팀 분석"), ("F-L-AT", "7. 해) 패=원정팀 분석"),
    ("F-WL-HT", "8. 해) 승/패=홈팀 분석"), ("F-WL-AT", "9. 해) 승/패=원정팀 분석"),
    ("TF-W", "10. 해/통) 승 분석"), ("TF-L", "11. 해/통) 패 분석"),
    ("TF-WL", "12. 해/통) 승+패 분석"), ("TF-WDL", "13. 해/통) 승+무+패 분석"),
    ("K-W", "14. 국) 승 분석"), ("K-L", "15. 국) 패 분석"), ("K-WL", "16. 국) 승+패 분석"),
    ("K-WDL", "17. 국) 승+무+패 분석"), ("K-W-HW", "18. 국) 승+H핸 분석"),
    ("K-W-HT", "19. 국) 승=홈팀 분석"), ("K-L-AT", "20. 국) 패=원정팀 분석"),
    ("K-WL-HT", "21. 국) 승/패=홈팀 분석"), ("K-WL-AT", "22. 국) 승/패=원정팀 분석"),
    ("TK-W", "23. 국/통) 승 분석"), ("TK-L", "24. 국/통) 패 분석"),
    ("TK-WL", "25. 국/통) 승+패 분석"), ("TK-WDL", "26. 국/통) 승+무+패 분석"),
]
_MYPICK_COLS = [
    ("IMPORTANT", "중요"), ("MY_PICK", "내픽"), ("MY_P", "P"), ("MY_HIT", "적중"), ("MEMO", "메모"),
]
# "P" 태그 배지 색 — 경기정보 그룹의 RT 배지와 같은 색 규칙(핸승=파랑/핸무=연파랑/무=회색/역=빨강).
_P_TAG_COLORS = {
    "핸승": {"bg": "1565C0", "fg": "FFFFFF", "bold": True},
    "핸무": {"bg": "64B5F6", "fg": "0D1B2A", "bold": True},
    "무": {"bg": "757575", "fg": "FFFFFF", "bold": True},
    "역": {"bg": "C62828", "fg": "FFFFFF", "bold": True},
}
_SUB4 = ["핸승", "핸무", "무", "역"]
_RT_DISPLAY = {1: "핸승", 2: "핸무", 3: "무", 4: "역", 5: "취소", 6: "연기"}
_RT_CODE_FROM_TEXT = {"핸승": 1, "핸무": 2, "무": 3, "역": 4, "취소": 5, "연기": 6}

_GROUP_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_GROUP_HEADER_FONT = Font(bold=True, color="FFFFFF")
_SUB_HEADER_FILL = PatternFill("solid", fgColor="374151")
_SUB_HEADER_FONT = Font(bold=True, color="E5E7EB", size=10)


def _blank(v):
    return v is None or v == ""


def _num(v):
    if _blank(v):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _rt_text(v):
    if _blank(v):
        return ""
    n = _num(v)
    if n is not None:
        return _RT_DISPLAY.get(int(n), "")
    return str(v)


def _rt_code(v):
    if _blank(v):
        return None
    n = _num(v)
    if n is not None:
        return int(n)
    return _RT_CODE_FROM_TEXT.get(str(v).strip())


def build_column_groups(available_cols):
    """columnGroups.js buildColumnGroups()와 동일. 실제로 존재하는 컬럼만 순서대로 배치."""
    available = set(available_cols)
    groups = []

    def add_flat(label1, label2, cols):
        leaves = [{"key": c, "sub": c} for c in cols if c in available]
        if leaves:
            groups.append({"label1": label1, "label2": label2, "kind": "flat", "cols": leaves})

    add_flat("일반정보", "시즌 및 라운드 정보", _GEN_COLS)
    add_flat("경기정보", "홈팀 vs 원정팀", _MATCH_COLS)
    add_flat("국내배당", "승(W) / 무(D) / 패(L)", _K_ODDS_COLS)
    add_flat("해외배당", "승(W) / 무(D) / 패(L)", _F_ODDS_COLS)

    mypick_leaves = [{"key": k, "sub": s} for k, s in _MYPICK_COLS if k in available]
    if mypick_leaves:
        groups.append({"label1": "내 예측", "label2": "", "kind": "mypick", "cols": mypick_leaves})

    ph_leaves = [{"key": k, "sub": s} for k, s in _PH_COLS if k in available]
    if ph_leaves:
        groups.append({"label1": "플핸 예측", "label2": "26개 지표 기반 · 실측 적중률",
                       "kind": "ph", "cols": ph_leaves})

    for code, title in _GROUP_DEFS:
        leaves = [{"key": f"{code} {i + 1}", "sub": sub} for i, sub in enumerate(_SUB4)
                 if f"{code} {i + 1}" in available]
        if leaves:
            groups.append({"label1": title, "label2": code, "kind": "indicator", "cols": leaves})

    return groups


def _format_cell(group, col, value):
    """columnGroups.js formatCell()과 동일한 표시 규칙."""
    g1, sub = group["label1"], col["sub"]

    if g1 == "경기정보" and sub == "RT":
        return _rt_text(value)
    if g1 == "경기정보" and sub in ("HS", "AS"):
        n = _num(value)
        return "" if n is None else str(int(n))
    if g1 == "일반정보" and sub == "No":
        n = _num(value)
        return "" if n is None else str(int(n))
    if g1 == "일반정보" and sub == "TM":
        n = _num(value)
        return "" if n is None else str(int(n)).zfill(4)
    if g1 in ("국내배당", "해외배당"):
        n = _num(value)
        if n is None:
            return ""
        if sub in ("KH", "FH"):
            return f"{'+' if n >= 0 else ''}{n:.0f}"
        return f"{n:.2f}"
    if group["kind"] == "ph":
        if sub == "PICK":
            return "" if _blank(value) else str(value)
        n = _num(value)
        return "" if n is None else f"{n:.0f}%"
    if group["kind"] == "mypick":
        if sub == "중요":
            return "★" if value is True or value == 1 or value == "1" else "☆"
        return "" if _blank(value) else str(value)
    if sub in _SUB4:
        n = _num(value)
        return "" if n is None else str(int(n))
    return "" if _blank(value) else str(value)


def _cell_style(group, col, value, row=None):
    """columnGroups.js cellStyle()과 동일한 배경/글자색 규칙. {bg, fg, bold} 또는 None.
    row는 HS/AS처럼 '이 행의 다른 컬럼 값'을 봐야 할 때만 쓴다(이긴 팀 점수 강조)."""
    g1, sub = group["label1"], col["sub"]

    if g1 == "경기정보" and sub in ("HS", "AS") and row is not None:
        hs, as_ = _num(row.get("HS")), _num(row.get("AS"))
        if hs is not None and as_ is not None and hs != as_:
            winner = "HS" if hs > as_ else "AS"
            if sub == winner:
                return {"fg": "C62828", "bold": True}
        return None

    if g1 == "경기정보" and sub == "RT":
        code = _rt_code(value)
        return {
            1: {"bg": "1565C0", "fg": "FFFFFF", "bold": True},
            2: {"bg": "64B5F6", "fg": "0D1B2A", "bold": True},
            3: {"bg": "757575", "fg": "FFFFFF", "bold": True},
            4: {"bg": "C62828", "fg": "FFFFFF", "bold": True},
            5: {"bg": "546E7A", "fg": "FFFFFF", "bold": True},
        }.get(code)

    if group["kind"] == "mypick" and sub == "적중":
        s = "" if _blank(value) else str(value).strip()
        return {
            "적중": {"bg": "FDD835", "fg": "0D1B2A", "bold": True},
            "미적": {"bg": "C62828", "fg": "FFFFFF", "bold": True},
            "패스": {"bg": "757575", "fg": "FFFFFF", "bold": True},
        }.get(s)

    if group["kind"] == "mypick" and sub == "P":
        s = "" if _blank(value) else str(value).strip()
        return _P_TAG_COLORS.get(s)

    if (g1 == "해외배당" and sub == "FH") or (g1 == "국내배당" and sub == "KH"):
        n = _num(value)
        if n is None:
            return None
        if n < 0:
            return {"fg": "1565C0", "bold": True}
        if n > 0:
            return {"fg": "C62828", "bold": True}
        return None

    if group["kind"] == "ph" and sub == "PICK":
        s = "" if _blank(value) else str(value).strip()
        if s.startswith("플핸"):
            if "(역)" in s:
                return {"bg": "4A148C", "fg": "FFFFFF", "bold": True}
            if "(무)" in s:
                return {"bg": "6A1B9A", "fg": "FFFFFF", "bold": True}
            if "(핸무)" in s:
                return {"bg": "E65100", "fg": "FFFFFF", "bold": True}
            return {"bg": "7B1FA2", "fg": "FFFFFF"}
        if s == "핸승":
            return {"bg": "1565C0", "fg": "FFFFFF", "bold": True}
        return {"fg": "9E9E9E"}

    if group["kind"] == "ph" and sub == "실측":
        n = _num(value)
        if n is None:
            return {"fg": "9E9E9E"}
        if n >= 80:
            return {"bg": "1B5E20", "fg": "FFFFFF", "bold": True}
        if n >= 75:
            return {"bg": "2E7D32", "fg": "FFFFFF", "bold": True}
        if n >= 70:
            return {"bg": "66BB6A", "fg": "0D1B2A"}
        if n >= 65:
            return {"bg": "C5E1A5", "fg": "1B5E20"}
        return {"fg": "9E9E9E"}

    if group["kind"] == "ph" and sub in ("해)플핸", "국)플핸"):
        n = _num(value)
        if n is None:
            return None
        if n >= 85:
            return {"bg": "311B92", "fg": "FFFFFF", "bold": True}
        if n >= 80:
            return {"bg": "512DA8", "fg": "FFFFFF"}
        if n >= 75:
            return {"bg": "9575CD", "fg": "1A1A1A"}
        if n < 50:
            return {"bg": "1565C0", "fg": "FFFFFF", "bold": True}
        return None

    return None


def build_table_excel(columns, rows, title: str = "분석표") -> io.BytesIO:
    """
    현재 조회된 분석표를 화면과 같은 2단 헤더(그룹제목+세부라벨) + 색상으로 내보낸다.
    (RT/PICK/실측/플핸% 배지 색, 핸디 +/- 색은 화면의 columnGroups.js cellStyle 규칙 그대로)
    """
    groups = build_column_groups(columns or [])

    wb = Workbook()
    ws = wb.active
    ws.title = (title or "분석표")[:31]

    leaf_meta = []  # 시트 컬럼 순서와 1:1 대응하는 (group, col)
    col_idx = 1
    for g in groups:
        span = len(g["cols"])
        start, end = col_idx, col_idx + span - 1

        head = ws.cell(row=1, column=start, value=g["label1"])
        head.font = _GROUP_HEADER_FONT
        head.alignment = _CENTER
        for c in range(start, end + 1):
            ws.cell(row=1, column=c).fill = _GROUP_HEADER_FILL
        if end > start:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)

        for i, col in enumerate(g["cols"]):
            sub_cell = ws.cell(row=2, column=start + i, value=col["sub"])
            sub_cell.font = _SUB_HEADER_FONT
            sub_cell.fill = _SUB_HEADER_FILL
            sub_cell.alignment = _CENTER
            leaf_meta.append((g, col))

        col_idx = end + 1

    for r_i, row in enumerate(rows or [], start=3):
        for c_i, (g, col) in enumerate(leaf_meta, start=1):
            value = row.get(col["key"])
            cell = ws.cell(row=r_i, column=c_i, value=_format_cell(g, col, value))
            style = _cell_style(g, col, value, row)
            if style:
                if "bg" in style:
                    cell.fill = PatternFill("solid", fgColor=style["bg"])
                cell.font = Font(bold=style.get("bold", False), color=style.get("fg"))

    ws.freeze_panes = "A3"
    for i, (g, col) in enumerate(leaf_meta, start=1):
        letter = get_column_letter(i)
        if col["sub"] in _SUB4:
            width = 7
        elif col["sub"] in ("DT",):
            width = 14
        elif col["sub"] in ("HT", "AT", "PICK"):
            width = 12
        else:
            width = 9
        ws.column_dimensions[letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
